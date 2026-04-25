



from datetime import datetime
import io
from itertools import groupby
import os
from bson import ObjectId
import openpyxl
from pydash import get, start_case, upper_case

from app.new_models.Discount import MemberType
from app.database.config import users
from app.new_models.Transaction import TransactionStatus
import json
import sys

def get_template_name(type: MemberType):
    return f'annex_{type.value}_template.xlsx'

def load_sheet(templateName: str):
    # templateName = get_template_name(type)
    fileName = os.path.join(os.getcwd(), 'app', 'templates', templateName)
    workbook = openpyxl.load_workbook(fileName)
    return workbook

def convert_to_bytes(workbook: openpyxl.Workbook):
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

def append_base_header(worksheet, user_id, data):

    branch = data[0]['branch'] if len(data) > 0 else None

    worksheet.cell(1, 1, "MMG-ALBAY")
    worksheet.cell(2, 1, upper_case(branch['streetAddress']))
    worksheet.cell(3, 1, 'VAT REG TIN ' + branch['tin'])
    worksheet.cell(9, 1, datetime.now().isoformat())
    worksheet.cell(5, 1, os.getenv('APP_VERSION'))

    user = users.find_one({ '_id': ObjectId(user_id) })
    worksheet.cell(10, 1, start_case(user['first_name'] + ' ' + user['last_name']))


def export_discount_reports(type: MemberType, reports, user_id):
    templateName = get_template_name(type)
    workbook = load_sheet(templateName)
    
    return export_discount_reports(workbook, reports, user_id)

def export_discount_reports(workbook: openpyxl.Workbook, type: MemberType, reports, user_id):
    # templateName = get_template_name(type)
    # workbook = load_sheet(templateName)
    worksheet = workbook.active

    append_base_header(worksheet, user_id, reports)

    if(type == MemberType.NAAC):
        append_naac_reports(worksheet, reports)
    elif(type == MemberType.SOLO_PARENT):
        append_solo_parent_reports(worksheet, reports)
    else:
        append_discount_reports(worksheet, reports)

    return convert_to_bytes(workbook)


def append_discount_reports(worksheet, reports):
    default_row = 17
    for index, report in enumerate(reports):
        default_col = 1
        worksheet.cell(column=default_col, row=default_row + index, value=report['transaction']['transactionDate'])
        worksheet.cell(column=default_col + 1, row=default_row + index, value=report['customer']['name'])
        worksheet.cell(column=default_col + 2, row=default_row + index, value=report['customer'].get('customer_type_id'))
        worksheet.cell(column=default_col + 3, row=default_row + index, value=report['customer']['tin_number'])
        worksheet.cell(column=default_col + 4, row=default_row + index, value=report['transaction']['invoiceNumber'])
        worksheet.cell(column=default_col + 7, row=default_row + index, value=report['transaction']['totalSalesWithoutMemberDiscount'])
        worksheet.cell(column=default_col + 9, row=default_row + index, value=report['transaction']['totalMemberDiscount'])
        worksheet.cell(column=default_col + 10, row=default_row + index, value=report['transaction']['totalNetSales'])

def append_naac_reports(worksheet, reports):
    default_row = 16
    for index, report in enumerate(reports):
        default_col = 1
        worksheet.cell(column=default_col, row=default_row + index, value=report['transaction']['transactionDate'])
        worksheet.cell(column=default_col + 1, row=default_row + index, value=report['customer']['name'])
        worksheet.cell(column=default_col + 2, row=default_row + index, value=report['customer'].get('customer_type_id'))
        worksheet.cell(column=default_col + 3, row=default_row + index, value=report['transaction']['invoiceNumber'])
        worksheet.cell(column=default_col + 4, row=default_row + index, value=report['transaction']['totalSalesWithoutMemberDiscount'])
        worksheet.cell(column=default_col + 5, row=default_row + index, value=report['transaction']['totalMemberDiscount'])
        worksheet.cell(column=default_col + 6, row=default_row + index, value=report['transaction']['totalNetSales'])

def append_solo_parent_reports(worksheet, reports):
    default_row = 17
    for index, report in enumerate(reports):
        default_col = 1
        worksheet.cell(column=default_col, row=default_row + index, value=report['transaction']['transactionDate'])
        worksheet.cell(column=default_col + 1, row=default_row + index, value=report['customer']['name'])
        worksheet.cell(column=default_col + 2, row=default_row + index, value=report['customer'].get('customer_type_id'))
        worksheet.cell(column=default_col + 6, row=default_row + index, value=report['transaction']['invoiceNumber'])
        worksheet.cell(column=default_col + 7, row=default_row + index, value=report['transaction']['totalSalesWithoutMemberDiscount'])
        worksheet.cell(column=default_col + 9, row=default_row + index, value=report['transaction']['totalMemberDiscount'])
        worksheet.cell(column=default_col + 10, row=default_row + index, value=report['transaction']['totalNetSales'])

def export_sales_reports(workbook, sales, user_id):
    worksheet = workbook.active
    
    append_base_header(worksheet, user_id, sales)
    append_sales_reports(worksheet, sales)
    return convert_to_bytes(workbook)

def append_sales_reports(worksheet, sales):
    default_row = 17
    def clip(value):
        return "{:.2f}".format(value)
    
    totalSales = sum(map(lambda i: i['totalNetSales'], sales))
    worksheet.cell(11, 1, f'Total: {clip(totalSales)}')


    for index, sale in enumerate(sales):
        default_col = 0
        col = default_col
        row = default_row + index

        worksheet.cell(row, col + 1, sale['date'])
        worksheet.cell(row, col + 2, str(sale['invoiceStartNumber']).zfill(6))
        worksheet.cell(row, col + 3, str(sale['invoiceEndNumber']).zfill(6))
        worksheet.cell(row, col + 4, clip(get(sale, 'endingCashCount.total', 0)))
        worksheet.cell(row, col + 5, clip(get(sale, 'openingFundd.total', 0)))
        worksheet.cell(row, col + 7, clip(sale['totalSalesWithoutMemberDiscount']))
        
        worksheet.cell(row, col + 10, clip(sale['totalSalesWithoutMemberDiscount']))

        discountSummary = sale['discountSummary']
        worksheet.cell(row, col + 12, clip(discountSummary.get(MemberType.SENIOR_CITIZEN.value, 0)))
        worksheet.cell(row, col + 13, clip(discountSummary.get(MemberType.PWD.value, 0)))
        worksheet.cell(row, col + 14, clip(discountSummary.get(MemberType.NAAC.value, 0)))
        worksheet.cell(row, col + 15, clip(discountSummary.get(MemberType.SOLO_PARENT.value, 0)))

        totalDiscount = sum(discountSummary.values())
        worksheet.cell(row, col + 19, clip(totalDiscount))

        worksheet.cell(row, col + 27, clip(sale['totalNetSales']))
        worksheet.cell(row, col + 28, clip(sale['cashDifference']))
        worksheet.cell(row, col + 30, 0)
        worksheet.cell(row, col + 31, 1)